"""
Excel Exporter
openpyxl ile Excel rapor oluşturma
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import Dict, List

from utils.logger import setup_logger

logger = setup_logger(__name__)


class ExcelExporter:
    """Excel export helper"""
    
    def __init__(self):
        self.header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.title_font = Font(bold=True, size=14, color="1976D2")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def export_patient_summary(self, report_data: Dict, output_path: str) -> bool:
        """Hasta özet raporu Excel"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Danışan Özeti"
            
            patient = report_data['patient']
            user = patient.user
            
            # Başlık
            ws['A1'] = f"Danışan Özet Raporu: {user.full_name if user else 'N/A'}"
            ws['A1'].font = self.title_font
            ws.merge_cells('A1:D1')
            
            # Hasta bilgileri
            row = 3
            ws[f'A{row}'] = "Hasta Bilgileri"
            ws[f'A{row}'].font = Font(bold=True, size=11)
            ws.merge_cells(f'A{row}:B{row}')
            
            row += 1
            info_data = [
                ['Email:', user.email if user else '-'],
                ['Yaş:', str(patient.age) if patient.age else '-'],
                ['Cinsiyet:', patient.gender or '-'],
                ['TC No:', patient.tc_no or '-'],
                ['Telefon:', user.phone if user and user.phone else '-']
            ]
            
            for label, value in info_data:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
            
            # İstatistikler
            row += 2
            ws[f'A{row}'] = "Görüşme İstatistikleri"
            ws[f'A{row}'].font = Font(bold=True, size=11)
            ws.merge_cells(f'A{row}:B{row}')
            
            row += 1
            stats_data = [
                ['Toplam Görüşme:', report_data['total_sessions']],
                ['Tamamlanan:', report_data['completed_sessions']],
                ['Son 30 Gün:', report_data['recent_sessions_count']],
                ['Toplam Süre (saat):', report_data['total_duration_hours']]
            ]
            
            for label, value in stats_data:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
            
            # Duygu analizi
            if report_data['emotion_summary']['total_analyzed'] > 0:
                row += 2
                ws[f'A{row}'] = "Duygu Analizi"
                ws[f'A{row}'].font = Font(bold=True, size=11)
                ws.merge_cells(f'A{row}:C{row}')
                
                row += 1
                ws[f'A{row}'] = "Duygu"
                ws[f'B{row}'] = "Sayı"
                ws[f'C{row}'] = "Yüzde"
                
                for col in ['A', 'B', 'C']:
                    ws[f'{col}{row}'].font = self.header_font
                    ws[f'{col}{row}'].fill = self.header_fill
                
                row += 1
                for emotion, count in report_data['emotion_summary']['distribution'].items():
                    ws[f'A{row}'] = emotion
                    ws[f'B{row}'] = count
                    ws[f'C{row}'] = f"{report_data['emotion_summary']['percentages'].get(emotion, 0)}%"
                    row += 1
            
            # Son görüşmeler
            row += 2
            ws[f'A{row}'] = "Son Görüşmeler"
            ws[f'A{row}'].font = Font(bold=True, size=11)
            ws.merge_cells(f'A{row}:D{row}')
            
            row += 1
            headers = ['Tarih', 'Süre (dk)', 'Durum', 'Notlar']
            for idx, header in enumerate(headers, start=1):
                col = get_column_letter(idx)
                ws[f'{col}{row}'] = header
                ws[f'{col}{row}'].font = self.header_font
                ws[f'{col}{row}'].fill = self.header_fill
            
            row += 1
            for session in report_data['sessions']:
                ws[f'A{row}'] = session.session_date.strftime('%d.%m.%Y') if session.session_date else '-'
                ws[f'B{row}'] = session.duration_minutes or '-'
                ws[f'C{row}'] = session.status
                ws[f'D{row}'] = (session.therapist_notes[:50] + '...' if session.therapist_notes and len(session.therapist_notes) > 50 else session.therapist_notes) or '-'
                row += 1
            
            # Sütun genişlikleri
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 40
            
            wb.save(output_path)
            logger.info(f"Excel created: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Excel export error: {e}")
            return False
    
    def export_period_report(self, report_data: Dict, output_path: str) -> bool:
        """Dönem raporu Excel"""
        try:
            wb = Workbook()
            
            # Özet sheet
            ws_summary = wb.active
            ws_summary.title = "Özet"
            
            period = report_data['period']
            ws_summary['A1'] = f"Dönem Raporu: {period['start'].strftime('%d.%m.%Y')} - {period['end'].strftime('%d.%m.%Y')}"
            ws_summary['A1'].font = self.title_font
            ws_summary.merge_cells('A1:C1')
            
            row = 3
            stats = report_data['statistics']
            stats_data = [
                ['Toplam Danışan:', stats['total_patients']],
                ['Aktif Danışan:', stats['active_patients']],
                ['Toplam Görüşme:', stats['total_sessions']],
                ['Tamamlanan Görüşme:', stats['completed_sessions']],
                ['İptal Edilen:', stats['cancelled_sessions']],
                ['Toplam Süre (saat):', stats['total_hours']],
                ['Ort. Görüşme Süresi (dk):', stats['avg_session_duration']]
            ]
            
            for label, value in stats_data:
                ws_summary[f'A{row}'] = label
                ws_summary[f'B{row}'] = value
                ws_summary[f'A{row}'].font = Font(bold=True)
                row += 1
            
            # Günlük dağılım sheet
            ws_daily = wb.create_sheet("Günlük Dağılım")
            ws_daily['A1'] = "Tarih"
            ws_daily['B1'] = "Görüşme Sayısı"
            ws_daily['A1'].font = self.header_font
            ws_daily['B1'].font = self.header_font
            ws_daily['A1'].fill = self.header_fill
            ws_daily['B1'].fill = self.header_fill
            
            row = 2
            for date, count in report_data['daily_distribution'].items():
                ws_daily[f'A{row}'] = date
                ws_daily[f'B{row}'] = count
                row += 1
            
            # Detaylı görüşmeler sheet
            ws_sessions = wb.create_sheet("Görüşmeler")
            headers = ['Tarih', 'Hasta', 'Süre (dk)', 'Durum']
            for idx, header in enumerate(headers, start=1):
                col = get_column_letter(idx)
                ws_sessions[f'{col}1'] = header
                ws_sessions[f'{col}1'].font = self.header_font
                ws_sessions[f'{col}1'].fill = self.header_fill
            
            row = 2
            for session in report_data['sessions']:
                ws_sessions[f'A{row}'] = session.session_date.strftime('%d.%m.%Y %H:%M') if session.session_date else '-'
                ws_sessions[f'B{row}'] = session.patient.user.full_name if session.patient and session.patient.user else '-'
                ws_sessions[f'C{row}'] = session.duration_minutes or '-'
                ws_sessions[f'D{row}'] = session.status
                row += 1
            
            # Sütun genişlikleri
            for ws in [ws_summary, ws_daily, ws_sessions]:
                for col in ['A', 'B', 'C', 'D']:
                    ws.column_dimensions[col].width = 20
            
            wb.save(output_path)
            logger.info(f"Period Excel created: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Period Excel export error: {e}")
            return False
    
    def export_session_list(self, sessions: List, output_path: str, title: str = "Görüşme Listesi") -> bool:
        """Görüşme listesi Excel"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Görüşmeler"
            
            ws['A1'] = title
            ws['A1'].font = self.title_font
            ws.merge_cells('A1:F1')
            
            # Headers
            headers = ['Tarih', 'Hasta', 'Süre (dk)', 'Durum', 'Tanı', 'Notlar']
            for idx, header in enumerate(headers, start=1):
                col = get_column_letter(idx)
                ws[f'{col}3'] = header
                ws[f'{col}3'].font = self.header_font
                ws[f'{col}3'].fill = self.header_fill
            
            row = 4
            for session in sessions:
                ws[f'A{row}'] = session.session_date.strftime('%d.%m.%Y %H:%M') if session.session_date else '-'
                ws[f'B{row}'] = session.patient.user.full_name if session.patient and session.patient.user else '-'
                ws[f'C{row}'] = session.duration_minutes or '-'
                ws[f'D{row}'] = session.status
                ws[f'E{row}'] = session.diagnosis or '-'
                ws[f'F{row}'] = (session.therapist_notes[:100] + '...' if session.therapist_notes and len(session.therapist_notes) > 100 else session.therapist_notes) or '-'
                row += 1
            
            # Sütun genişlikleri
            ws.column_dimensions['A'].width = 18
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 50
            
            wb.save(output_path)
            logger.info(f"Session list Excel created: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Session list export error: {e}")
            return False